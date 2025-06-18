import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_final_enhanced_master():
    """Create the ultimate enhanced Excel file with fund details for all analyses"""
    
    print("Creating FINAL Enhanced Master Analysis Excel File...")
    print("=" * 60)
    
    # Load all data
    q1_2025 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q1_2025.csv')
    q4_2024 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q4_2024.csv')
    
    print(f"Loaded Q1 2025: {len(q1_2025)} holdings")
    print(f"Loaded Q4 2024: {len(q4_2024)} holdings")
    
    # Perform all analyses with fund details
    overall_analysis = perform_enhanced_overall_analysis(q1_2025)
    position_changes = perform_enhanced_position_changes_analysis(q1_2025, q4_2024)
    
    # Create comprehensive Excel workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="2E5894")
    subtitle_font = Font(bold=True, size=12, color="2E5894")
    
    # Color scheme
    summary_fill = PatternFill(start_color="2E5894", end_color="2E5894", fill_type="solid")      # Dark Blue
    frequent_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")     # Green  
    value_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")       # Blue
    new_buy_fill = PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")     # Light Green
    increase_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")    # Amber
    exit_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")        # Red
    
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create all sheets
    create_executive_summary(wb, q1_2025, q4_2024, overall_analysis, position_changes)
    create_enhanced_frequent_holdings_sheet(wb, overall_analysis, frequent_fill, header_font, header_alignment, border)
    create_enhanced_top_values_sheet(wb, overall_analysis, value_fill, header_font, header_alignment, border)
    create_enhanced_weight_analysis_sheet(wb, overall_analysis, value_fill, header_font, header_alignment, border)
    create_enhanced_new_positions_sheet(wb, position_changes['new_buys'], new_buy_fill, header_font, header_alignment, border)
    create_enhanced_increases_sheet(wb, position_changes['increases'], increase_fill, header_font, header_alignment, border)
    create_enhanced_exits_sheet(wb, position_changes['exits'], exit_fill, header_font, header_alignment, border)
    create_raw_data_sheet(wb, q1_2025, "Q1 2025 Raw Data", header_font, summary_fill, header_alignment)
    
    # Save the master file
    filename = "FINAL ENHANCED - Biotech 13F Complete Analysis Q1 2025.xlsx"
    wb.save(filename)
    
    print(f"\n✅ FINAL ENHANCED MASTER ANALYSIS COMPLETE!")
    print(f"📊 Created: {filename}")
    print(f"📋 Contains 8 comprehensive analysis sheets with complete fund details")
    
    return filename

def perform_enhanced_overall_analysis(df):
    """Perform enhanced overall holdings analysis with fund tracking"""
    
    # Create mapping of companies to funds holding them
    company_to_funds = defaultdict(lambda: {'funds': set(), 'total_value': 0, 'positions': []})
    
    for _, holding in df.iterrows():
        company_key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        fund_name = holding['fund_name']
        
        company_to_funds[company_key]['funds'].add(fund_name)
        company_to_funds[company_key]['total_value'] += holding['value']
        company_to_funds[company_key]['positions'].append({
            'fund': fund_name,
            'value': holding['value'],
            'shares': holding['shares'],
            'ticker': holding['ticker'],
            'industry': holding['industry']
        })
    
    # Most frequently held stocks with fund details
    frequent_holdings = []
    for company, data in company_to_funds.items():
        fund_count = len(data['funds'])
        fund_list = sorted(list(data['funds']))
        
        # Get representative data
        sample_position = data['positions'][0] if data['positions'] else {}
        
        frequent_holdings.append({
            'company': company,
            'ticker': sample_position.get('ticker', ''),
            'fund_count': fund_count,
            'fund_list': fund_list,
            'fund_names': ', '.join(fund_list),
            'total_value': data['total_value'],
            'avg_position_size': data['total_value'] / fund_count if fund_count > 0 else 0,
            'industry': sample_position.get('industry', '')
        })
    
    # Sort by frequency then by total value
    frequent_holdings.sort(key=lambda x: (x['fund_count'], x['total_value']), reverse=True)
    top_frequent = frequent_holdings[:25]
    
    # Top holdings by total value with fund details
    value_holdings = sorted(frequent_holdings, key=lambda x: x['total_value'], reverse=True)[:25]
    
    # Portfolio weight analysis
    total_value_all = df['value'].sum()
    weight_holdings = []
    
    for holding in value_holdings:
        weight_percent = (holding['total_value'] / total_value_all) * 100
        consensus_score = weight_percent * holding['fund_count']
        
        weight_holdings.append({
            'company': holding['company'],
            'ticker': holding['ticker'],
            'fund_count': holding['fund_count'],
            'fund_list': holding['fund_list'],
            'fund_names': holding['fund_names'],
            'total_value': holding['total_value'],
            'weight_percent': weight_percent,
            'consensus_score': consensus_score,
            'avg_position_size': holding['avg_position_size'],
            'industry': holding['industry']
        })
    
    return {
        'top_frequent': top_frequent,
        'top_values': value_holdings,
        'weight_analysis': weight_holdings,
        'total_value': total_value_all
    }

def perform_enhanced_position_changes_analysis(q1_2025, q4_2024):
    """Enhanced position changes analysis with fund tracking"""
    
    def create_position_key(row):
        return f"{row['fund_name']}||{row['company']}"
    
    # Create position dictionaries
    q1_positions = {}
    q4_positions = {}
    
    for _, row in q1_2025.iterrows():
        key = create_position_key(row)
        q1_positions[key] = {
            'value': row['value'],
            'shares': row['shares'],
            'company': row['company'],
            'ticker': row['ticker'],
            'fund_name': row['fund_name'],
            'industry': row['industry']
        }
    
    for _, row in q4_2024.iterrows():
        key = create_position_key(row)
        q4_positions[key] = {
            'value': row['value'],
            'shares': row['shares'],
            'company': row['company'],
            'ticker': row['ticker'],
            'fund_name': row['fund_name'],
            'industry': row['industry']
        }
    
    # Analyze changes with fund tracking
    new_buys = []
    position_exits = []
    position_increases = []
    
    # New buys
    for key in q1_positions:
        if key not in q4_positions:
            new_buys.append(q1_positions[key])
    
    # Exits
    for key in q4_positions:
        if key not in q1_positions:
            position_exits.append(q4_positions[key])
    
    # Increases
    for key in q1_positions:
        if key in q4_positions:
            q1_value = q1_positions[key]['value']
            q4_value = q4_positions[key]['value']
            
            if q4_value > 0:
                pct_change = (q1_value - q4_value) / q4_value
                
                if pct_change > 0.5:  # >50% increase
                    increase_data = q1_positions[key].copy()
                    increase_data['q4_value'] = q4_value
                    increase_data['q1_value'] = q1_value
                    increase_data['pct_change'] = pct_change
                    increase_data['dollar_change'] = q1_value - q4_value
                    position_increases.append(increase_data)
    
    # Create enhanced summaries with fund names
    def create_enhanced_summary(positions, category_name):
        company_to_funds = defaultdict(lambda: {'funds': [], 'positions': []})
        
        # Group by company and collect fund info
        for pos in positions:
            company = pos['company']
            fund = pos['fund_name']
            company_to_funds[company]['funds'].append(fund)
            company_to_funds[company]['positions'].append(pos)
        
        # Create detailed analysis
        detailed = []
        for company, data in company_to_funds.items():
            funds = data['funds']
            positions = data['positions']
            fund_count = len(funds)
            
            total_value = sum(pos['value'] for pos in positions)
            sample_pos = positions[0] if positions else {}
            
            detail = {
                'company': company,
                'ticker': sample_pos.get('ticker', ''),
                'funds_count': fund_count,
                'fund_names': ', '.join(sorted(funds)),
                'total_value': total_value,
                'avg_position_size': total_value / fund_count if fund_count > 0 else 0,
                'industry': sample_pos.get('industry', '')
            }
            
            # Add specific metrics for increases
            if category_name == 'increases':
                if any('pct_change' in pos for pos in positions):
                    avg_pct_change = sum(pos.get('pct_change', 0) for pos in positions) / len(positions)
                    total_dollar_change = sum(pos.get('dollar_change', 0) for pos in positions)
                    detail['avg_pct_increase'] = avg_pct_change
                    detail['total_dollar_increase'] = total_dollar_change
            
            detailed.append(detail)
        
        # Sort by fund count then by total value
        detailed.sort(key=lambda x: (x['funds_count'], x['total_value']), reverse=True)
        return detailed[:25]  # Top 25
    
    return {
        'new_buys': create_enhanced_summary(new_buys, 'new_buys'),
        'increases': create_enhanced_summary(position_increases, 'increases'),
        'exits': create_enhanced_summary(position_exits, 'exits')
    }

def create_enhanced_new_positions_sheet(wb, new_buys, fill_color, header_font, header_alignment, border):
    """Create enhanced new positions sheet with fund names"""
    
    ws = wb.create_sheet("🆕 New Positions")
    
    headers = ["Rank", "Company", "Ticker", "Funds Buying", "Fund Names", "Total Value ($)", "Avg Position ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, buy in enumerate(new_buys, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=buy['company']).border = border
        ws.cell(row=i, column=3, value=buy['ticker']).border = border
        ws.cell(row=i, column=4, value=buy['funds_count']).border = border
        ws.cell(row=i, column=5, value=buy['fund_names']).border = border
        ws.cell(row=i, column=6, value=buy['total_value']).border = border
        ws.cell(row=i, column=7, value=buy['avg_position_size']).border = border
        ws.cell(row=i, column=8, value=buy['industry']).border = border
    
    # Auto-adjust columns with wider Fund Names column
    for col_idx, column in enumerate(ws.columns, 1):
        if col_idx == 5:  # Fund Names column
            ws.column_dimensions[column[0].column_letter].width = 80  # Extra wide for fund names
        else:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_enhanced_increases_sheet(wb, increases, fill_color, header_font, header_alignment, border):
    """Create enhanced position increases sheet with fund names"""
    
    ws = wb.create_sheet("📊 Position Increases")
    
    headers = ["Rank", "Company", "Ticker", "Funds Adding", "Fund Names", "Current Value ($)", "$ Increase", "Avg % Increase", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, inc in enumerate(increases, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=inc['company']).border = border
        ws.cell(row=i, column=3, value=inc['ticker']).border = border
        ws.cell(row=i, column=4, value=inc['funds_count']).border = border
        ws.cell(row=i, column=5, value=inc['fund_names']).border = border
        ws.cell(row=i, column=6, value=inc['total_value']).border = border
        ws.cell(row=i, column=7, value=inc.get('total_dollar_increase', 0)).border = border
        avg_pct = inc.get('avg_pct_increase', 0)
        ws.cell(row=i, column=8, value=f"{avg_pct:.1%}").border = border
        ws.cell(row=i, column=9, value=inc['industry']).border = border
    
    # Auto-adjust columns with wider Fund Names column
    for col_idx, column in enumerate(ws.columns, 1):
        if col_idx == 5:  # Fund Names column
            ws.column_dimensions[column[0].column_letter].width = 80  # Extra wide for fund names
        else:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_enhanced_exits_sheet(wb, exits, fill_color, header_font, header_alignment, border):
    """Create enhanced position exits sheet with fund names"""
    
    ws = wb.create_sheet("📉 Position Exits")
    
    headers = ["Rank", "Company", "Ticker", "Funds Exiting", "Fund Names", "Exit Value ($)", "Avg Position ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, exit in enumerate(exits, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=exit['company']).border = border
        ws.cell(row=i, column=3, value=exit['ticker']).border = border
        ws.cell(row=i, column=4, value=exit['funds_count']).border = border
        ws.cell(row=i, column=5, value=exit['fund_names']).border = border
        ws.cell(row=i, column=6, value=exit['total_value']).border = border
        ws.cell(row=i, column=7, value=exit['avg_position_size']).border = border
        ws.cell(row=i, column=8, value=exit['industry']).border = border
    
    # Auto-adjust columns with wider Fund Names column
    for col_idx, column in enumerate(ws.columns, 1):
        if col_idx == 5:  # Fund Names column
            ws.column_dimensions[column[0].column_letter].width = 80  # Extra wide for fund names
        else:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

# Reuse functions from previous versions
def create_enhanced_frequent_holdings_sheet(wb, analysis, fill_color, header_font, header_alignment, border):
    """Create enhanced most frequently held sheet with fund names"""
    
    ws = wb.create_sheet("🏆 Most Frequently Held")
    
    headers = ["Rank", "Company", "Ticker", "Funds Holding", "Fund Names", "Total Value ($)", "Avg Position ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, holding in enumerate(analysis['top_frequent'], 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=holding['company']).border = border
        ws.cell(row=i, column=3, value=holding['ticker']).border = border
        ws.cell(row=i, column=4, value=holding['fund_count']).border = border
        ws.cell(row=i, column=5, value=holding['fund_names']).border = border
        ws.cell(row=i, column=6, value=holding['total_value']).border = border
        ws.cell(row=i, column=7, value=holding['avg_position_size']).border = border
        ws.cell(row=i, column=8, value=holding['industry']).border = border
    
    # Auto-adjust columns with wider Fund Names column
    for col_idx, column in enumerate(ws.columns, 1):
        if col_idx == 5:  # Fund Names column
            ws.column_dimensions[column[0].column_letter].width = 80  # Extra wide for fund names
        else:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_enhanced_top_values_sheet(wb, analysis, fill_color, header_font, header_alignment, border):
    """Create enhanced top holdings by value sheet with fund names"""
    
    ws = wb.create_sheet("💰 Top Holdings by Value")
    
    headers = ["Rank", "Company", "Ticker", "Total Value ($)", "Portfolio Weight (%)", "Funds Holding", "Fund Names", "Avg Position ($)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, holding in enumerate(analysis['top_values'], 2):
        weight_percent = (holding['total_value'] / analysis['total_value']) * 100
        
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=holding['company']).border = border
        ws.cell(row=i, column=3, value=holding['ticker']).border = border
        ws.cell(row=i, column=4, value=holding['total_value']).border = border
        ws.cell(row=i, column=5, value=f"{weight_percent:.2f}%").border = border
        ws.cell(row=i, column=6, value=holding['fund_count']).border = border
        ws.cell(row=i, column=7, value=holding['fund_names']).border = border
        ws.cell(row=i, column=8, value=holding['avg_position_size']).border = border
    
    # Auto-adjust columns with wider Fund Names column
    for col_idx, column in enumerate(ws.columns, 1):
        if col_idx == 7:  # Fund Names column
            ws.column_dimensions[column[0].column_letter].width = 80  # Extra wide for fund names
        else:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_enhanced_weight_analysis_sheet(wb, analysis, fill_color, header_font, header_alignment, border):
    """Create enhanced portfolio weight analysis sheet with fund names"""
    
    ws = wb.create_sheet("📈 Portfolio Weight Analysis")
    
    headers = ["Rank", "Company", "Ticker", "Portfolio Weight (%)", "Consensus Score", "Funds Holding", "Fund Names", "Total Value ($)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    # Sort by consensus score (weight * fund count)
    sorted_holdings = sorted(analysis['weight_analysis'], key=lambda x: x['consensus_score'], reverse=True)
    
    for i, holding in enumerate(sorted_holdings, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=holding['company']).border = border
        ws.cell(row=i, column=3, value=holding['ticker']).border = border
        ws.cell(row=i, column=4, value=f"{holding['weight_percent']:.2f}%").border = border
        ws.cell(row=i, column=5, value=f"{holding['consensus_score']:.1f}").border = border
        ws.cell(row=i, column=6, value=holding['fund_count']).border = border
        ws.cell(row=i, column=7, value=holding['fund_names']).border = border
        ws.cell(row=i, column=8, value=holding['total_value']).border = border
    
    # Auto-adjust columns with wider Fund Names column
    for col_idx, column in enumerate(ws.columns, 1):
        if col_idx == 7:  # Fund Names column
            ws.column_dimensions[column[0].column_letter].width = 80  # Extra wide for fund names
        else:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_executive_summary(wb, q1_2025, q4_2024, overall_analysis, position_changes):
    """Create executive summary sheet"""
    
    ws = wb.active
    ws.title = "📊 Executive Summary"
    
    title_font = Font(bold=True, size=16, color="2E5894")
    subtitle_font = Font(bold=True, size=12, color="2E5894")
    
    # Main title
    ws['A1'] = "BIOTECH HEDGE FUND 13F COMPLETE ANALYSIS (FINAL ENHANCED)"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "Q1 2025 Holdings & Q4→Q1 Position Changes with Complete Fund Details"
    ws['A2'].font = Font(bold=True, size=12, color="666666")
    ws.merge_cells('A2:D2')
    
    # Q1 2025 Portfolio Overview
    row = 4
    ws[f'A{row}'] = "Q1 2025 PORTFOLIO OVERVIEW"
    ws[f'A{row}'].font = subtitle_font
    row += 1
    
    portfolio_stats = [
        ("Total Funds Analyzed:", len(q1_2025['fund_name'].unique())),
        ("Total Holdings:", len(q1_2025)),
        ("Unique Companies:", len(q1_2025['company'].unique())),
        ("Total Portfolio Value:", f"${overall_analysis['total_value']:,.0f}"),
        ("Average Position Size:", f"${q1_2025['value'].mean():,.0f}"),
        ("Largest Single Position:", f"${q1_2025['value'].max():,.0f}"),
    ]
    
    for label, value in portfolio_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Position Changes Overview
    row += 2
    ws[f'A{row}'] = "POSITION CHANGES: Q4 2024 → Q1 2025"
    ws[f'A{row}'].font = subtitle_font
    row += 1
    
    changes_stats = [
        ("New Positions Initiated:", len([b for b in position_changes['new_buys'] if b['funds_count'] > 0])),
        ("Position Increases (>50%):", len([i for i in position_changes['increases'] if i['funds_count'] > 0])),
        ("Position Exits:", len([e for e in position_changes['exits'] if e['funds_count'] > 0])),
        ("New Buy Total Value:", f"${sum(b['total_value'] for b in position_changes['new_buys']):,.0f}"),
        ("Exit Total Value:", f"${sum(e['total_value'] for e in position_changes['exits']):,.0f}"),
    ]
    
    for label, value in changes_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Final enhancement note
    row += 2
    ws[f'A{row}'] = "✨ FINAL ENHANCED FEATURES"
    ws[f'A{row}'].font = subtitle_font
    row += 1
    
    enhancements = [
        "📋 Fund names for ALL analysis categories",
        "🆕 New positions with fund details", 
        "📈 Position increases with fund details",
        "📉 Position exits with fund details",
        "🎯 Complete transparency on all fund moves",
        "🔍 Ultimate smart money tracking capabilities",
    ]
    
    for enhancement in enhancements:
        ws[f'A{row}'] = enhancement
        row += 1

def create_raw_data_sheet(wb, df, sheet_name, header_font, fill_color, header_alignment):
    """Create raw data sheet (limited to first 1000 rows)"""
    
    ws = wb.create_sheet(sheet_name)
    
    # Add data (limit to 1000 rows for Excel performance)
    for r_idx, row in enumerate(dataframe_to_rows(df.head(1000), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = fill_color
                cell.alignment = header_alignment

if __name__ == "__main__":
    print("CREATING FINAL ENHANCED MASTER ANALYSIS WITH COMPLETE FUND DETAILS")
    print("=" * 75)
    
    filename = create_final_enhanced_master()
    
    print(f"\n🎯 SUCCESS! Created ultimate comprehensive analysis file:")
    print(f"📁 {filename}")
    print(f"\n📋 Final enhanced sheets with complete fund details:")
    print(f"   1. 📊 Executive Summary")
    print(f"   2. 🏆 Most Frequently Held (with fund names)")
    print(f"   3. 💰 Top Holdings by Value (with fund names)")
    print(f"   4. 📈 Portfolio Weight Analysis (with fund names)")
    print(f"   5. 🆕 New Positions (with fund names)")
    print(f"   6. 📊 Position Increases (with fund names)")
    print(f"   7. 📉 Position Exits (with fund names)")
    print(f"   8. Q1 2025 Raw Data")
    print(f"\n✨ ULTIMATE ENHANCEMENT: Fund names for ALL position changes!")
    print(f"🎯 Now see exactly which funds made every move!")