import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def analyze_new_buys_sells():
    """Compare Q1 2025 vs Q4 2024 to identify new buys and sells"""
    
    print("Loading Q1 2025 and Q4 2024 data...")
    
    # Load the data
    q1_2025 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q1_2025.csv')
    q4_2024 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q4_2024.csv')
    
    print(f"Q1 2025: {len(q1_2025)} holdings from {len(q1_2025['fund_name'].unique())} funds")
    print(f"Q4 2024: {len(q4_2024)} holdings from {len(q4_2024['fund_name'].unique())} funds")
    
    # Create position keys for comparison (fund + company)
    def create_position_key(row):
        return f"{row['fund_name']}||{row['company']}"
    
    # Create sets of positions for each quarter
    q1_positions = set(q1_2025.apply(create_position_key, axis=1))
    q4_positions = set(q4_2024.apply(create_position_key, axis=1))
    
    # Find new buys (in Q1 but not in Q4)
    new_buys = q1_positions - q4_positions
    
    # Find sells/exits (in Q4 but not in Q1)
    new_sells = q4_positions - q1_positions
    
    print(f"\nIdentified {len(new_buys)} new positions (buys)")
    print(f"Identified {len(new_sells)} position exits (sells)")
    
    # Analyze new buys
    new_buy_companies = Counter()
    new_buy_details = []
    
    for position_key in new_buys:
        fund_name, company = position_key.split('||')
        new_buy_companies[company] += 1
        
        # Get details from Q1 data
        q1_row = q1_2025[(q1_2025['fund_name'] == fund_name) & (q1_2025['company'] == company)]
        if not q1_row.empty:
            row = q1_row.iloc[0]
            new_buy_details.append({
                'company': company,
                'ticker': row['ticker'],
                'fund_name': fund_name,
                'value': row['value'],
                'shares': row['shares'],
                'industry': row['industry']
            })
    
    # Analyze sells
    new_sell_companies = Counter()
    new_sell_details = []
    
    for position_key in new_sells:
        fund_name, company = position_key.split('||')
        new_sell_companies[company] += 1
        
        # Get details from Q4 data
        q4_row = q4_2024[(q4_2024['fund_name'] == fund_name) & (q4_2024['company'] == company)]
        if not q4_row.empty:
            row = q4_row.iloc[0]
            new_sell_details.append({
                'company': company,
                'ticker': row['ticker'],
                'fund_name': fund_name,
                'value': row['value'],
                'shares': row['shares'],
                'industry': row['industry']
            })
    
    # Get top 30 most common new buys
    top_30_buys = new_buy_companies.most_common(30)
    
    # Get top 30 most common sells
    top_30_sells = new_sell_companies.most_common(30)
    
    print(f"\n🔥 TOP 30 MOST COMMON NEW BUYS:")
    print("-" * 60)
    for i, (company, count) in enumerate(top_30_buys, 1):
        print(f"{i:2d}. {company}: {count} funds")
    
    print(f"\n📉 TOP 30 MOST COMMON SELLS/EXITS:")
    print("-" * 60)
    for i, (company, count) in enumerate(top_30_sells, 1):
        print(f"{i:2d}. {company}: {count} funds")
    
    # Calculate total values for new buys and sells
    buy_value_totals = defaultdict(float)
    sell_value_totals = defaultdict(float)
    
    for detail in new_buy_details:
        buy_value_totals[detail['company']] += detail['value']
    
    for detail in new_sell_details:
        sell_value_totals[detail['company']] += detail['value']
    
    # Create detailed analysis
    top_buys_with_details = []
    for company, count in top_30_buys:
        total_value = buy_value_totals[company]
        # Get representative ticker and industry
        sample_detail = next((d for d in new_buy_details if d['company'] == company), {})
        ticker = sample_detail.get('ticker', '')
        industry = sample_detail.get('industry', '')
        
        top_buys_with_details.append({
            'company': company,
            'ticker': ticker,
            'funds_buying': count,
            'total_value': total_value,
            'avg_position_size': total_value / count if count > 0 else 0,
            'industry': industry
        })
    
    top_sells_with_details = []
    for company, count in top_30_sells:
        total_value = sell_value_totals[company]
        # Get representative ticker and industry
        sample_detail = next((d for d in new_sell_details if d['company'] == company), {})
        ticker = sample_detail.get('ticker', '')
        industry = sample_detail.get('industry', '')
        
        top_sells_with_details.append({
            'company': company,
            'ticker': ticker,
            'funds_selling': count,
            'total_value': total_value,
            'avg_position_size': total_value / count if count > 0 else 0,
            'industry': industry
        })
    
    # Create Excel report
    create_buys_sells_excel(top_buys_with_details, top_sells_with_details, new_buy_details, new_sell_details)
    
    # Create CSV files
    create_buys_sells_csv(top_buys_with_details, top_sells_with_details)
    
    return top_buys_with_details, top_sells_with_details

def create_buys_sells_excel(top_buys, top_sells, buy_details, sell_details):
    """Create comprehensive Excel report for new buys and sells"""
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    buy_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
    sell_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")  # Red
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    # Title
    ws_summary['A1'] = "Q1 2025 vs Q4 2024 NEW BUYS & SELLS ANALYSIS"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    # Summary stats
    row = 3
    total_new_buys = len(buy_details)
    total_new_sells = len(sell_details)
    total_buy_value = sum(detail['value'] for detail in buy_details)
    total_sell_value = sum(detail['value'] for detail in sell_details)
    
    summary_data = [
        ("Analysis Period:", "Q4 2024 → Q1 2025"),
        ("Total New Positions:", f"{total_new_buys:,}"),
        ("Total Position Exits:", f"{total_new_sells:,}"),
        ("Total New Buy Value:", f"${total_buy_value:,.0f}"),
        ("Total Sell Value:", f"${total_sell_value:,.0f}"),
        ("Net Activity:", f"${total_buy_value - total_sell_value:,.0f}"),
        ("Unique Companies Bought:", len(set(detail['company'] for detail in buy_details))),
        ("Unique Companies Sold:", len(set(detail['company'] for detail in sell_details)))
    ]
    
    for label, value in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Sheet 2: Top 30 New Buys
    ws_buys = wb.create_sheet("Top 30 New Buys")
    
    # Headers
    headers = ["Rank", "Company", "Ticker", "Funds Buying", "Total Value ($)", "Avg Position Size ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws_buys.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = buy_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    for i, buy in enumerate(top_buys, 2):
        ws_buys.cell(row=i, column=1, value=i-1).border = border
        ws_buys.cell(row=i, column=2, value=buy['company']).border = border
        ws_buys.cell(row=i, column=3, value=buy['ticker']).border = border
        ws_buys.cell(row=i, column=4, value=buy['funds_buying']).border = border
        ws_buys.cell(row=i, column=5, value=buy['total_value']).border = border
        ws_buys.cell(row=i, column=6, value=buy['avg_position_size']).border = border
        ws_buys.cell(row=i, column=7, value=buy['industry']).border = border
    
    # Auto-adjust column widths
    for column in ws_buys.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_buys.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 3: Top 30 Sells
    ws_sells = wb.create_sheet("Top 30 Sells")
    
    # Headers
    headers = ["Rank", "Company", "Ticker", "Funds Selling", "Total Value ($)", "Avg Position Size ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws_sells.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = sell_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    for i, sell in enumerate(top_sells, 2):
        ws_sells.cell(row=i, column=1, value=i-1).border = border
        ws_sells.cell(row=i, column=2, value=sell['company']).border = border
        ws_sells.cell(row=i, column=3, value=sell['ticker']).border = border
        ws_sells.cell(row=i, column=4, value=sell['funds_selling']).border = border
        ws_sells.cell(row=i, column=5, value=sell['total_value']).border = border
        ws_sells.cell(row=i, column=6, value=sell['avg_position_size']).border = border
        ws_sells.cell(row=i, column=7, value=sell['industry']).border = border
    
    # Auto-adjust column widths
    for column in ws_sells.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_sells.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Save
    filename = "Top New Buys Sells Q1 2025.xlsx"
    wb.save(filename)
    print(f"\n✅ Excel report saved: {filename}")

def create_buys_sells_csv(top_buys, top_sells):
    """Create CSV files for new buys and sells"""
    
    # Create DataFrames
    buys_df = pd.DataFrame(top_buys)
    sells_df = pd.DataFrame(top_sells)
    
    # Add rank column
    buys_df.insert(0, 'rank', range(1, len(buys_df) + 1))
    sells_df.insert(0, 'rank', range(1, len(sells_df) + 1))
    
    # Save CSV files
    buys_df.to_csv("Top New Buys Q1 2025.csv", index=False)
    sells_df.to_csv("Top New Sells Q1 2025.csv", index=False)
    
    print(f"✅ CSV files saved:")
    print(f"   - Top New Buys Q1 2025.csv")
    print(f"   - Top New Sells Q1 2025.csv")

if __name__ == "__main__":
    print("ANALYZING NEW BUYS & SELLS: Q4 2024 → Q1 2025")
    print("=" * 60)
    
    top_buys, top_sells = analyze_new_buys_sells()
    
    print(f"\n🎉 ANALYSIS COMPLETE!")
    print(f"📊 Files generated:")
    print(f"   - Top New Buys Sells Q1 2025.xlsx")
    print(f"   - Top New Buys Q1 2025.csv")
    print(f"   - Top New Sells Q1 2025.csv")