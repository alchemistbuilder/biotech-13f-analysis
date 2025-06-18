import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def analyze_buys_sells_increases():
    """Compare Q1 2025 vs Q4 2024 to identify new buys, sells, and position increases"""
    
    print("Loading Q1 2025 and Q4 2024 data...")
    
    # Load the data
    q1_2025 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q1_2025.csv')
    q4_2024 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q4_2024.csv')
    
    print(f"Q1 2025: {len(q1_2025)} holdings from {len(q1_2025['fund_name'].unique())} funds")
    print(f"Q4 2024: {len(q4_2024)} holdings from {len(q4_2024['fund_name'].unique())} funds")
    
    # Create position keys for comparison (fund + company)
    def create_position_key(row):
        return f"{row['fund_name']}||{row['company']}"
    
    # Create dictionaries for position values by fund-company key
    q1_positions = {}
    q4_positions = {}
    
    for _, row in q1_2025.iterrows():
        key = create_position_key(row)
        q1_positions[key] = {
            'value': row['value'],
            'shares': row['shares'],
            'company': row['company'],
            'ticker': row['ticker'],
            'fund_name': row['fund_name'],
            'industry': row['industry']
        }
    
    for _, row in q4_2024.iterrows():
        key = create_position_key(row)
        q4_positions[key] = {
            'value': row['value'],
            'shares': row['shares'],
            'company': row['company'],
            'ticker': row['ticker'],
            'fund_name': row['fund_name'],
            'industry': row['industry']
        }
    
    # Find different types of changes
    new_buys = []           # Positions in Q1 but not in Q4
    position_exits = []     # Positions in Q4 but not in Q1
    position_increases = [] # Positions that increased significantly
    position_decreases = [] # Positions that decreased significantly
    
    # Analyze new buys
    for key in q1_positions:
        if key not in q4_positions:
            new_buys.append(q1_positions[key])
    
    # Analyze exits
    for key in q4_positions:
        if key not in q1_positions:
            position_exits.append(q4_positions[key])
    
    # Analyze position changes (existing positions)
    for key in q1_positions:
        if key in q4_positions:
            q1_value = q1_positions[key]['value']
            q4_value = q4_positions[key]['value']
            
            if q4_value > 0:  # Avoid division by zero
                pct_change = (q1_value - q4_value) / q4_value
                
                # Significant increase (>50% increase)
                if pct_change > 0.5:
                    increase_data = q1_positions[key].copy()
                    increase_data['q4_value'] = q4_value
                    increase_data['q1_value'] = q1_value
                    increase_data['pct_change'] = pct_change
                    increase_data['dollar_change'] = q1_value - q4_value
                    position_increases.append(increase_data)
                
                # Significant decrease (>50% decrease)
                elif pct_change < -0.5:
                    decrease_data = q4_positions[key].copy()
                    decrease_data['q4_value'] = q4_value
                    decrease_data['q1_value'] = q1_value
                    decrease_data['pct_change'] = pct_change
                    decrease_data['dollar_change'] = q1_value - q4_value
                    position_decreases.append(decrease_data)
    
    print(f"\nIdentified {len(new_buys)} completely new positions")
    print(f"Identified {len(position_exits)} position exits") 
    print(f"Identified {len(position_increases)} significant position increases (>50%)")
    print(f"Identified {len(position_decreases)} significant position decreases (>50%)")
    
    # Count by company for new buys
    new_buy_companies = Counter(pos['company'] for pos in new_buys)
    
    # Count by company for exits
    exit_companies = Counter(pos['company'] for pos in position_exits)
    
    # Count by company for increases
    increase_companies = Counter(pos['company'] for pos in position_increases)
    
    # Count by company for decreases
    decrease_companies = Counter(pos['company'] for pos in position_decreases)
    
    # Get top 30 for each category
    top_30_new_buys = new_buy_companies.most_common(30)
    top_30_exits = exit_companies.most_common(30)
    top_30_increases = increase_companies.most_common(30)
    top_30_decreases = decrease_companies.most_common(30)
    
    print(f"\n🔥 TOP 30 COMPLETELY NEW POSITIONS:")
    print("-" * 60)
    for i, (company, count) in enumerate(top_30_new_buys, 1):
        print(f"{i:2d}. {company}: {count} funds")
    
    print(f"\n📈 TOP 30 POSITION INCREASES (>50% increase):")
    print("-" * 60) 
    for i, (company, count) in enumerate(top_30_increases, 1):
        print(f"{i:2d}. {company}: {count} funds doubled down")
    
    print(f"\n📉 TOP 30 POSITION EXITS:")
    print("-" * 60)
    for i, (company, count) in enumerate(top_30_exits, 1):
        print(f"{i:2d}. {company}: {count} funds exited")
    
    # Create detailed analysis with values
    def create_detailed_analysis(positions, company_counter, category_name):
        """Create detailed analysis with total values and averages"""
        detailed = []
        
        for company, count in company_counter.most_common(30):
            company_positions = [pos for pos in positions if pos['company'] == company]
            
            total_value = sum(pos['value'] for pos in company_positions)
            avg_value = total_value / len(company_positions) if company_positions else 0
            
            # Get representative data
            sample_pos = company_positions[0] if company_positions else {}
            ticker = sample_pos.get('ticker', '')
            industry = sample_pos.get('industry', '')
            
            detail = {
                'company': company,
                'ticker': ticker,
                'funds_count': count,
                'total_value': total_value,
                'avg_position_size': avg_value,
                'industry': industry
            }
            
            # Add specific metrics for increases
            if category_name == 'increases' and 'pct_change' in sample_pos:
                avg_pct_change = sum(pos['pct_change'] for pos in company_positions) / len(company_positions)
                total_dollar_change = sum(pos['dollar_change'] for pos in company_positions)
                detail['avg_pct_increase'] = avg_pct_change
                detail['total_dollar_increase'] = total_dollar_change
            
            detailed.append(detail)
        
        return detailed
    
    # Create detailed analyses
    new_buys_detailed = create_detailed_analysis(new_buys, new_buy_companies, 'new_buys')
    increases_detailed = create_detailed_analysis(position_increases, increase_companies, 'increases')
    exits_detailed = create_detailed_analysis(position_exits, exit_companies, 'exits')
    decreases_detailed = create_detailed_analysis(position_decreases, decrease_companies, 'decreases')
    
    # Create Excel report
    create_comprehensive_excel(new_buys_detailed, increases_detailed, exits_detailed, decreases_detailed)
    
    # Create CSV files
    create_comprehensive_csv(new_buys_detailed, increases_detailed, exits_detailed, decreases_detailed)
    
    return new_buys_detailed, increases_detailed, exits_detailed, decreases_detailed

def create_comprehensive_excel(new_buys, increases, exits, decreases):
    """Create comprehensive Excel report"""
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    new_buy_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
    increase_fill = PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")  # Light Green
    exit_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")     # Red
    decrease_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid") # Orange
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    ws_summary['A1'] = "Q1 2025 COMPREHENSIVE POSITION CHANGES ANALYSIS"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    # Summary stats
    row = 3
    summary_data = [
        ("Analysis Period:", "Q4 2024 → Q1 2025"),
        ("", ""),
        ("NEW POSITIONS:", ""),
        ("  Total New Positions:", len([b for b in new_buys if b['funds_count'] > 0])),
        ("  Total New Buy Value:", f"${sum(b['total_value'] for b in new_buys):,.0f}"),
        ("", ""),
        ("POSITION INCREASES (>50%):", ""),
        ("  Companies with Increases:", len([i for i in increases if i['funds_count'] > 0])),
        ("  Total Increase Value:", f"${sum(i.get('total_dollar_increase', i['total_value']) for i in increases):,.0f}"),
        ("", ""),
        ("POSITION EXITS:", ""),
        ("  Total Position Exits:", len([e for e in exits if e['funds_count'] > 0])),
        ("  Total Exit Value:", f"${sum(e['total_value'] for e in exits):,.0f}"),
    ]
    
    for label, value in summary_data:
        ws_summary[f'A{row}'] = label
        if label and not label.startswith(" "):
            ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Sheet 2: New Positions
    ws_new = wb.create_sheet("New Positions")
    
    headers = ["Rank", "Company", "Ticker", "Funds Buying", "Total Value ($)", "Avg Position Size ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = new_buy_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for i, buy in enumerate(new_buys, 2):
        ws_new.cell(row=i, column=1, value=i-1).border = border
        ws_new.cell(row=i, column=2, value=buy['company']).border = border
        ws_new.cell(row=i, column=3, value=buy['ticker']).border = border
        ws_new.cell(row=i, column=4, value=buy['funds_count']).border = border
        ws_new.cell(row=i, column=5, value=buy['total_value']).border = border
        ws_new.cell(row=i, column=6, value=buy['avg_position_size']).border = border
        ws_new.cell(row=i, column=7, value=buy['industry']).border = border
    
    # Auto-adjust column widths
    for column in ws_new.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_new.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 3: Position Increases
    ws_inc = wb.create_sheet("Position Increases")
    
    headers = ["Rank", "Company", "Ticker", "Funds Adding", "Total Current Value ($)", "Total $ Increase", "Avg % Increase", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws_inc.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = increase_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for i, inc in enumerate(increases, 2):
        ws_inc.cell(row=i, column=1, value=i-1).border = border
        ws_inc.cell(row=i, column=2, value=inc['company']).border = border
        ws_inc.cell(row=i, column=3, value=inc['ticker']).border = border
        ws_inc.cell(row=i, column=4, value=inc['funds_count']).border = border
        ws_inc.cell(row=i, column=5, value=inc['total_value']).border = border
        ws_inc.cell(row=i, column=6, value=inc.get('total_dollar_increase', 0)).border = border
        avg_pct = inc.get('avg_pct_increase', 0)
        ws_inc.cell(row=i, column=7, value=f"{avg_pct:.1%}").border = border
        ws_inc.cell(row=i, column=8, value=inc['industry']).border = border
    
    # Auto-adjust column widths
    for column in ws_inc.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_inc.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 4: Position Exits
    ws_exits = wb.create_sheet("Position Exits")
    
    headers = ["Rank", "Company", "Ticker", "Funds Exiting", "Total Exit Value ($)", "Avg Position Size ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws_exits.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = exit_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for i, exit in enumerate(exits, 2):
        ws_exits.cell(row=i, column=1, value=i-1).border = border
        ws_exits.cell(row=i, column=2, value=exit['company']).border = border
        ws_exits.cell(row=i, column=3, value=exit['ticker']).border = border
        ws_exits.cell(row=i, column=4, value=exit['funds_count']).border = border
        ws_exits.cell(row=i, column=5, value=exit['total_value']).border = border
        ws_exits.cell(row=i, column=6, value=exit['avg_position_size']).border = border
        ws_exits.cell(row=i, column=7, value=exit['industry']).border = border
    
    # Auto-adjust column widths
    for column in ws_exits.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_exits.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Save
    filename = "Top New Buys Increases Sells Q1 2025.xlsx"
    wb.save(filename)
    print(f"\n✅ Comprehensive Excel report saved: {filename}")

def create_comprehensive_csv(new_buys, increases, exits, decreases):
    """Create CSV files for all categories"""
    
    # New positions
    new_df = pd.DataFrame(new_buys)
    if not new_df.empty:
        new_df.insert(0, 'rank', range(1, len(new_df) + 1))
        new_df.to_csv("Top New Positions Q1 2025.csv", index=False)
    
    # Position increases
    inc_df = pd.DataFrame(increases)
    if not inc_df.empty:
        inc_df.insert(0, 'rank', range(1, len(inc_df) + 1))
        inc_df.to_csv("Top Position Increases Q1 2025.csv", index=False)
    
    # Position exits
    exit_df = pd.DataFrame(exits)
    if not exit_df.empty:
        exit_df.insert(0, 'rank', range(1, len(exit_df) + 1))
        exit_df.to_csv("Top Position Exits Q1 2025.csv", index=False)
    
    print(f"✅ CSV files saved:")
    print(f"   - Top New Positions Q1 2025.csv")
    print(f"   - Top Position Increases Q1 2025.csv") 
    print(f"   - Top Position Exits Q1 2025.csv")

if __name__ == "__main__":
    print("COMPREHENSIVE POSITION ANALYSIS: Q4 2024 → Q1 2025")
    print("=" * 65)
    
    new_buys, increases, exits, decreases = analyze_buys_sells_increases()
    
    print(f"\n🎉 COMPREHENSIVE ANALYSIS COMPLETE!")
    print(f"📊 Files generated:")
    print(f"   - Top New Buys Increases Sells Q1 2025.xlsx")
    print(f"   - Top New Positions Q1 2025.csv")
    print(f"   - Top Position Increases Q1 2025.csv")
    print(f"   - Top Position Exits Q1 2025.csv")