import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import Counter, defaultdict
import locale

def create_biotech_13f_analysis():
    """Create the final 1Q 2025 Biotech 13F Analysis report"""
    
    # Read the combined data
    df = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q1_2025.csv')
    
    print(f"Creating 1Q 2025 Biotech 13F Analysis...")
    print(f"Processing {len(df)} holdings from {len(df['fund_name'].unique())} funds")
    
    # Analysis 1: Top 25 most frequently held stocks
    stock_frequency = Counter()
    for _, holding in df.iterrows():
        key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        stock_frequency[key] += 1
    
    top_frequent = stock_frequency.most_common(25)
    
    # Analysis 2: Top 25 holdings by total value
    value_totals = defaultdict(float)
    for _, holding in df.iterrows():
        key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        value_totals[key] += holding['value']
    
    top_values = sorted(value_totals.items(), key=lambda x: x[1], reverse=True)[:25]
    
    # Analysis 3: Top 25 holdings by percentage weight
    total_value_all = df['value'].sum()
    weight_analysis = []
    for stock, total_value in top_values:
        weight_percent = (total_value / total_value_all) * 100
        weight_analysis.append((stock, weight_percent, total_value))
    
    # Create Excel workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E5894", end_color="2E5894", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14, color="2E5894")
    subtitle_font = Font(bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    # Title
    ws_summary['A1'] = "1Q 2025 BIOTECH HEDGE FUND 13F ANALYSIS"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    # Summary stats
    row = 3
    summary_data = [
        ("Total Funds Analyzed:", len(df['fund_name'].unique())),
        ("Total Holdings:", len(df)),
        ("Unique Companies:", len(df['company'].unique())),
        ("Total Portfolio Value:", f"${total_value_all:,.0f}"),
        ("Average Position Size:", f"${df['value'].mean():,.0f}"),
        ("Reporting Period:", "Q1 2025 (March 31, 2025)")
    ]
    
    for label, value in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = subtitle_font
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Funds list
    row += 2
    ws_summary[f'A{row}'] = "FUNDS ANALYZED:"
    ws_summary[f'A{row}'].font = subtitle_font
    row += 1
    
    for i, fund_name in enumerate(sorted(df['fund_name'].unique()), 1):
        ws_summary[f'A{row}'] = f"{i}. {fund_name}"
        row += 1
    
    # Sheet 2: Most Frequently Held
    ws_freq = wb.create_sheet("Most Frequently Held")
    
    # Headers
    headers = ["Rank", "Company", "Number of Funds Holding"]
    for col, header in enumerate(headers, 1):
        cell = ws_freq.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    for i, (stock, count) in enumerate(top_frequent, 2):
        ws_freq.cell(row=i, column=1, value=i-1).border = border
        ws_freq.cell(row=i, column=2, value=stock).border = border
        ws_freq.cell(row=i, column=3, value=count).border = border
    
    # Auto-adjust column widths
    for column in ws_freq.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_freq.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Sheet 3: Top Holdings by Value
    ws_value = wb.create_sheet("Top Holdings by Value")
    
    # Headers
    headers = ["Rank", "Company", "Total Value ($)", "Market Share (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws_value.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    for i, (stock, weight, value) in enumerate(weight_analysis, 2):
        ws_value.cell(row=i, column=1, value=i-1).border = border
        ws_value.cell(row=i, column=2, value=stock).border = border
        ws_value.cell(row=i, column=3, value=value).border = border
        ws_value.cell(row=i, column=4, value=f"{weight:.2f}%").border = border
    
    # Auto-adjust column widths
    for column in ws_value.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_value.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Sheet 4: Portfolio Weights
    ws_weight = wb.create_sheet("Portfolio Weight Analysis")
    
    # Headers
    headers = ["Rank", "Company", "Portfolio Weight (%)", "Total Value ($)", "Funds Holding"]
    for col, header in enumerate(headers, 1):
        cell = ws_weight.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data with fund count
    for i, (stock, weight, value) in enumerate(weight_analysis, 2):
        fund_count = stock_frequency[stock]
        ws_weight.cell(row=i, column=1, value=i-1).border = border
        ws_weight.cell(row=i, column=2, value=stock).border = border
        ws_weight.cell(row=i, column=3, value=f"{weight:.2f}%").border = border
        ws_weight.cell(row=i, column=4, value=value).border = border
        ws_weight.cell(row=i, column=5, value=fund_count).border = border
    
    # Auto-adjust column widths
    for column in ws_weight.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_weight.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Sheet 5: Raw Data (first 1000 rows for Excel limits)
    ws_raw = wb.create_sheet("Raw Holdings Data")
    
    # Add headers with styling
    for r_idx, row in enumerate(dataframe_to_rows(df.head(1000), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    # Save the file
    filename = "1Q 2025 Biotech 13F Analysis.xlsx"
    wb.save(filename)
    
    # Also create CSV versions
    freq_df = pd.DataFrame(top_frequent, columns=['Company', 'Funds_Holding'])
    freq_df.index += 1
    freq_df.to_csv("1Q 2025 Biotech 13F Analysis - Most Frequent.csv")
    
    value_df = pd.DataFrame([(stock, value, weight) for stock, weight, value in weight_analysis], 
                           columns=['Company', 'Total_Value', 'Portfolio_Weight_Percent'])
    value_df.index += 1
    value_df.to_csv("1Q 2025 Biotech 13F Analysis - Top Values.csv")
    
    print(f"\n✅ ANALYSIS COMPLETE!")
    print(f"📊 Files created:")
    print(f"   - {filename}")
    print(f"   - 1Q 2025 Biotech 13F Analysis - Most Frequent.csv")
    print(f"   - 1Q 2025 Biotech 13F Analysis - Top Values.csv")
    
    # Print top 25 summaries
    print(f"\n🏆 TOP 25 MOST FREQUENTLY HELD STOCKS:")
    print("-" * 60)
    for i, (stock, count) in enumerate(top_frequent, 1):
        print(f"{i:2d}. {stock}: {count} funds")
    
    print(f"\n💰 TOP 25 HOLDINGS BY TOTAL VALUE:")
    print("-" * 60)
    for i, (stock, weight, value) in enumerate(weight_analysis, 1):
        print(f"{i:2d}. {stock}: ${value:,.0f} ({weight:.2f}%)")
    
    return filename

if __name__ == "__main__":
    create_biotech_13f_analysis()