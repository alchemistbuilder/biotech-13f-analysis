import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_master_analysis():
    """Create one comprehensive Excel file with all analyses"""
    
    print("Creating Master Analysis Excel File...")
    print("=" * 50)
    
    # Load all data
    q1_2025 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q1_2025.csv')
    q4_2024 = pd.read_csv('ALL_HEDGE_FUNDS_13F_Q4_2024.csv')
    
    print(f"Loaded Q1 2025: {len(q1_2025)} holdings")
    print(f"Loaded Q4 2024: {len(q4_2024)} holdings")
    
    # Perform all analyses
    overall_analysis = perform_overall_analysis(q1_2025)
    position_changes = perform_position_changes_analysis(q1_2025, q4_2024)
    
    # Create comprehensive Excel workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="2E5894")
    subtitle_font = Font(bold=True, size=12, color="2E5894")
    
    # Color scheme
    summary_fill = PatternFill(start_color="2E5894", end_color="2E5894", fill_type="solid")      # Dark Blue
    frequent_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")     # Green  
    value_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")       # Blue
    new_buy_fill = PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")     # Light Green
    increase_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")    # Amber
    exit_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")        # Red
    
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Executive Summary
    create_executive_summary(wb, q1_2025, q4_2024, overall_analysis, position_changes)
    
    # Sheet 2: Most Frequently Held
    create_frequent_holdings_sheet(wb, overall_analysis['top_frequent'], frequent_fill, header_font, header_alignment, border)
    
    # Sheet 3: Top Holdings by Value 
    create_top_values_sheet(wb, overall_analysis['top_values'], overall_analysis['weight_analysis'], value_fill, header_font, header_alignment, border)
    
    # Sheet 4: Portfolio Weight Analysis
    create_weight_analysis_sheet(wb, overall_analysis['weight_analysis'], overall_analysis['stock_frequency'], value_fill, header_font, header_alignment, border)
    
    # Sheet 5: New Positions
    create_new_positions_sheet(wb, position_changes['new_buys'], new_buy_fill, header_font, header_alignment, border)
    
    # Sheet 6: Position Increases
    create_increases_sheet(wb, position_changes['increases'], increase_fill, header_font, header_alignment, border)
    
    # Sheet 7: Position Exits
    create_exits_sheet(wb, position_changes['exits'], exit_fill, header_font, header_alignment, border)
    
    # Sheet 8: Raw Data Q1 2025 (sample)
    create_raw_data_sheet(wb, q1_2025, "Q1 2025 Raw Data", header_font, summary_fill, header_alignment)
    
    # Save the master file
    filename = "MASTER - Biotech 13F Complete Analysis Q1 2025.xlsx"
    wb.save(filename)
    
    print(f"\n✅ MASTER ANALYSIS COMPLETE!")
    print(f"📊 Created: {filename}")
    print(f"📋 Contains 8 comprehensive analysis sheets")
    
    return filename

def perform_overall_analysis(df):
    """Perform the overall holdings analysis"""
    
    # Most frequently held stocks
    stock_frequency = Counter()
    for _, holding in df.iterrows():
        key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        stock_frequency[key] += 1
    
    top_frequent = stock_frequency.most_common(25)
    
    # Top holdings by total value
    value_totals = defaultdict(float)
    for _, holding in df.iterrows():
        key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        value_totals[key] += holding['value']
    
    top_values = sorted(value_totals.items(), key=lambda x: x[1], reverse=True)[:25]
    
    # Portfolio weight analysis
    total_value_all = df['value'].sum()
    weight_analysis = []
    for stock, total_value in top_values:
        weight_percent = (total_value / total_value_all) * 100
        weight_analysis.append((stock, weight_percent, total_value))
    
    return {
        'top_frequent': top_frequent,
        'top_values': top_values,
        'weight_analysis': weight_analysis,
        'stock_frequency': stock_frequency,
        'total_value': total_value_all
    }

def perform_position_changes_analysis(q1_2025, q4_2024):
    """Perform position changes analysis"""
    
    def create_position_key(row):
        return f"{row['fund_name']}||{row['company']}"
    
    # Create position dictionaries
    q1_positions = {}
    q4_positions = {}
    
    for _, row in q1_2025.iterrows():
        key = create_position_key(row)
        q1_positions[key] = {
            'value': row['value'],
            'shares': row['shares'],
            'company': row['company'],
            'ticker': row['ticker'],
            'fund_name': row['fund_name'],
            'industry': row['industry']
        }
    
    for _, row in q4_2024.iterrows():
        key = create_position_key(row)
        q4_positions[key] = {
            'value': row['value'],
            'shares': row['shares'],
            'company': row['company'],
            'ticker': row['ticker'],
            'fund_name': row['fund_name'],
            'industry': row['industry']
        }
    
    # Analyze changes
    new_buys = []
    position_exits = []
    position_increases = []
    
    # New buys
    for key in q1_positions:
        if key not in q4_positions:
            new_buys.append(q1_positions[key])
    
    # Exits
    for key in q4_positions:
        if key not in q1_positions:
            position_exits.append(q4_positions[key])
    
    # Increases
    for key in q1_positions:
        if key in q4_positions:
            q1_value = q1_positions[key]['value']
            q4_value = q4_positions[key]['value']
            
            if q4_value > 0:
                pct_change = (q1_value - q4_value) / q4_value
                
                if pct_change > 0.5:  # >50% increase
                    increase_data = q1_positions[key].copy()
                    increase_data['q4_value'] = q4_value
                    increase_data['q1_value'] = q1_value
                    increase_data['pct_change'] = pct_change
                    increase_data['dollar_change'] = q1_value - q4_value
                    position_increases.append(increase_data)
    
    # Create summaries
    def create_summary(positions, category_name):
        company_counter = Counter(pos['company'] for pos in positions)
        detailed = []
        
        for company, count in company_counter.most_common(25):
            company_positions = [pos for pos in positions if pos['company'] == company]
            
            total_value = sum(pos['value'] for pos in company_positions)
            sample_pos = company_positions[0] if company_positions else {}
            
            detail = {
                'company': company,
                'ticker': sample_pos.get('ticker', ''),
                'funds_count': count,
                'total_value': total_value,
                'avg_position_size': total_value / count if count > 0 else 0,
                'industry': sample_pos.get('industry', '')
            }
            
            if category_name == 'increases' and 'pct_change' in sample_pos:
                avg_pct_change = sum(pos['pct_change'] for pos in company_positions) / len(company_positions)
                total_dollar_change = sum(pos['dollar_change'] for pos in company_positions)
                detail['avg_pct_increase'] = avg_pct_change
                detail['total_dollar_increase'] = total_dollar_change
            
            detailed.append(detail)
        
        return detailed
    
    return {
        'new_buys': create_summary(new_buys, 'new_buys'),
        'increases': create_summary(position_increases, 'increases'),
        'exits': create_summary(position_exits, 'exits')
    }

def create_executive_summary(wb, q1_2025, q4_2024, overall_analysis, position_changes):
    """Create executive summary sheet"""
    
    ws = wb.active
    ws.title = "📊 Executive Summary"
    
    title_font = Font(bold=True, size=16, color="2E5894")
    subtitle_font = Font(bold=True, size=12, color="2E5894")
    
    # Main title
    ws['A1'] = "BIOTECH HEDGE FUND 13F COMPLETE ANALYSIS"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "Q1 2025 Holdings & Q4→Q1 Position Changes"
    ws['A2'].font = Font(bold=True, size=12, color="666666")
    ws.merge_cells('A2:D2')
    
    # Q1 2025 Portfolio Overview
    row = 4
    ws[f'A{row}'] = "Q1 2025 PORTFOLIO OVERVIEW"
    ws[f'A{row}'].font = subtitle_font
    row += 1
    
    portfolio_stats = [
        ("Total Funds Analyzed:", len(q1_2025['fund_name'].unique())),
        ("Total Holdings:", len(q1_2025)),
        ("Unique Companies:", len(q1_2025['company'].unique())),
        ("Total Portfolio Value:", f"${overall_analysis['total_value']:,.0f}"),
        ("Average Position Size:", f"${q1_2025['value'].mean():,.0f}"),
        ("Largest Single Position:", f"${q1_2025['value'].max():,.0f}"),
    ]
    
    for label, value in portfolio_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Position Changes Overview
    row += 2
    ws[f'A{row}'] = "POSITION CHANGES: Q4 2024 → Q1 2025"
    ws[f'A{row}'].font = subtitle_font
    row += 1
    
    changes_stats = [
        ("New Positions Initiated:", len([b for b in position_changes['new_buys'] if b['funds_count'] > 0])),
        ("Position Increases (>50%):", len([i for i in position_changes['increases'] if i['funds_count'] > 0])),
        ("Position Exits:", len([e for e in position_changes['exits'] if e['funds_count'] > 0])),
        ("New Buy Total Value:", f"${sum(b['total_value'] for b in position_changes['new_buys']):,.0f}"),
        ("Exit Total Value:", f"${sum(e['total_value'] for e in position_changes['exits']):,.0f}"),
    ]
    
    for label, value in changes_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Top themes
    row += 2
    ws[f'A{row}'] = "TOP INVESTMENT THEMES"
    ws[f'A{row}'].font = subtitle_font
    row += 1
    
    themes = [
        "🧬 Gene Therapy & Cell Therapy",
        "💊 Rare Disease Treatments", 
        "🧠 CNS & Neurological Disorders",
        "🔬 Precision Medicine & Biomarkers",
        "⚕️ Immunology & Autoimmune",
    ]
    
    for theme in themes:
        ws[f'A{row}'] = theme
        row += 1

def create_frequent_holdings_sheet(wb, top_frequent, fill_color, header_font, header_alignment, border):
    """Create most frequently held sheet"""
    
    ws = wb.create_sheet("🏆 Most Frequently Held")
    
    headers = ["Rank", "Company", "Number of Funds Holding"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, (stock, count) in enumerate(top_frequent, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=stock).border = border
        ws.cell(row=i, column=3, value=count).border = border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)

def create_top_values_sheet(wb, top_values, weight_analysis, fill_color, header_font, header_alignment, border):
    """Create top holdings by value sheet"""
    
    ws = wb.create_sheet("💰 Top Holdings by Value")
    
    headers = ["Rank", "Company", "Total Value ($)", "Portfolio Weight (%)", "Market Share"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, (stock, weight, value) in enumerate(weight_analysis, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=stock).border = border
        ws.cell(row=i, column=3, value=value).border = border
        ws.cell(row=i, column=4, value=f"{weight:.2f}%").border = border
        ws.cell(row=i, column=5, value=f"${value:,.0f}").border = border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)

def create_weight_analysis_sheet(wb, weight_analysis, stock_frequency, fill_color, header_font, header_alignment, border):
    """Create portfolio weight analysis sheet"""
    
    ws = wb.create_sheet("📈 Portfolio Weight Analysis")
    
    headers = ["Rank", "Company", "Portfolio Weight (%)", "Total Value ($)", "Funds Holding", "Consensus Score"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, (stock, weight, value) in enumerate(weight_analysis, 2):
        fund_count = stock_frequency[stock]
        consensus_score = weight * fund_count  # Weight × Fund Count
        
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=stock).border = border
        ws.cell(row=i, column=3, value=f"{weight:.2f}%").border = border
        ws.cell(row=i, column=4, value=value).border = border
        ws.cell(row=i, column=5, value=fund_count).border = border
        ws.cell(row=i, column=6, value=f"{consensus_score:.1f}").border = border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)

def create_new_positions_sheet(wb, new_buys, fill_color, header_font, header_alignment, border):
    """Create new positions sheet"""
    
    ws = wb.create_sheet("🆕 New Positions")
    
    headers = ["Rank", "Company", "Ticker", "Funds Buying", "Total Value ($)", "Avg Position ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, buy in enumerate(new_buys, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=buy['company']).border = border
        ws.cell(row=i, column=3, value=buy['ticker']).border = border
        ws.cell(row=i, column=4, value=buy['funds_count']).border = border
        ws.cell(row=i, column=5, value=buy['total_value']).border = border
        ws.cell(row=i, column=6, value=buy['avg_position_size']).border = border
        ws.cell(row=i, column=7, value=buy['industry']).border = border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_increases_sheet(wb, increases, fill_color, header_font, header_alignment, border):
    """Create position increases sheet"""
    
    ws = wb.create_sheet("📊 Position Increases")
    
    headers = ["Rank", "Company", "Ticker", "Funds Adding", "Current Value ($)", "$ Increase", "Avg % Increase", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, inc in enumerate(increases, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=inc['company']).border = border
        ws.cell(row=i, column=3, value=inc['ticker']).border = border
        ws.cell(row=i, column=4, value=inc['funds_count']).border = border
        ws.cell(row=i, column=5, value=inc['total_value']).border = border
        ws.cell(row=i, column=6, value=inc.get('total_dollar_increase', 0)).border = border
        avg_pct = inc.get('avg_pct_increase', 0)
        ws.cell(row=i, column=7, value=f"{avg_pct:.1%}").border = border
        ws.cell(row=i, column=8, value=inc['industry']).border = border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_exits_sheet(wb, exits, fill_color, header_font, header_alignment, border):
    """Create position exits sheet"""
    
    ws = wb.create_sheet("📉 Position Exits")
    
    headers = ["Rank", "Company", "Ticker", "Funds Exiting", "Exit Value ($)", "Avg Position ($)", "Industry"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = fill_color
        cell.alignment = header_alignment
        cell.border = border
    
    for i, exit in enumerate(exits, 2):
        ws.cell(row=i, column=1, value=i-1).border = border
        ws.cell(row=i, column=2, value=exit['company']).border = border
        ws.cell(row=i, column=3, value=exit['ticker']).border = border
        ws.cell(row=i, column=4, value=exit['funds_count']).border = border
        ws.cell(row=i, column=5, value=exit['total_value']).border = border
        ws.cell(row=i, column=6, value=exit['avg_position_size']).border = border
        ws.cell(row=i, column=7, value=exit['industry']).border = border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def create_raw_data_sheet(wb, df, sheet_name, header_font, fill_color, header_alignment):
    """Create raw data sheet (limited to first 1000 rows)"""
    
    ws = wb.create_sheet(sheet_name)
    
    # Add data (limit to 1000 rows for Excel performance)
    for r_idx, row in enumerate(dataframe_to_rows(df.head(1000), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = fill_color
                cell.alignment = header_alignment

if __name__ == "__main__":
    print("CREATING MASTER ANALYSIS FILE")
    print("=" * 40)
    
    filename = create_master_analysis()
    
    print(f"\n🎯 SUCCESS! Created comprehensive analysis file:")
    print(f"📁 {filename}")
    print(f"\n📋 Contains the following sheets:")
    print(f"   1. 📊 Executive Summary")
    print(f"   2. 🏆 Most Frequently Held")
    print(f"   3. 💰 Top Holdings by Value")
    print(f"   4. 📈 Portfolio Weight Analysis")
    print(f"   5. 🆕 New Positions")
    print(f"   6. 📊 Position Increases")
    print(f"   7. 📉 Position Exits")
    print(f"   8. Q1 2025 Raw Data")