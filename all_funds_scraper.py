import requests
import pandas as pd
import time

def get_all_hedge_funds_13f():
    """Get Q1 2025 13F data for all 16 hedge funds"""
    
    api_key = "m1HzjYgss43pOkJZcWTr2tuKvRvOPM4W"
    date = "2025-03-31"  # Q1 2025
    
    # All 16 hedge funds with their CIKs
    hedge_funds = {
        "Avoro Capital Advisors LLC": "0001633313",
        "Baker Bros. Advisors LP": "0001263508", 
        "BVF Inc": "0001056807",
        "Checkpoint Capital L.P.": "0001977548",
        "Commodore Capital LP": "0001831942",
        "Cormorant Asset Management, LP": "0001583977",
        "Darwin Global Management, Ltd.": "0001839209",
        "Frazier Life Sciences Management, L.P.": "0001892134",
        "Logos Global Management LP": "0001792126",
        "Lynx1 Capital Management LP": "0001910456", 
        "Paradigm Biocapital Advisors LP": "0001855655",
        "Perceptive Advisors LLC": "0001224962",
        "Ra Capital Management, L.P.": "0001346824",
        "Rock Springs Capital Management LP": "0001595725",
        "Rtw Investments, LP": "0001493215",
        "Vivo Capital, LLC": "0001674712"
    }
    
    print(f"Getting Q1 2025 13F data for all {len(hedge_funds)} hedge funds...")
    print("="*70)
    
    all_holdings = []
    successful_funds = []
    
    for fund_name, cik in hedge_funds.items():
        try:
            print(f"\n📊 Processing {fund_name} (CIK: {cik})...")
            
            url = f"https://financialmodelingprep.com/api/v4/institutional-ownership/portfolio-holdings?cik={cik}&date={date}&apikey={api_key}"
            
            response = requests.get(url, timeout=20)
            print(f"   Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                
                if isinstance(data, list) and len(data) > 0:
                    print(f"   ✅ SUCCESS! Found {len(data)} holdings")
                    
                    # Process holdings for this fund
                    fund_holdings = []
                    total_value = 0
                    
                    for item in data:
                        if isinstance(item, dict):
                            holding = {
                                'fund_name': fund_name,
                                'cik': cik,
                                'company': item.get('securityName', ''),
                                'ticker': item.get('symbol', ''),
                                'cusip': item.get('securityCusip', ''),
                                'shares': item.get('sharesNumber', 0),
                                'value': item.get('marketValue', 0),
                                'weight': item.get('weight', 0),
                                'ownership': item.get('ownership', 0),
                                'date': date,
                                'filing_date': item.get('filingDate', ''),
                                'industry': item.get('industryTitle', '')
                            }
                            fund_holdings.append(holding)
                            total_value += holding['value']
                    
                    # Sort by value
                    fund_holdings.sort(key=lambda x: x['value'], reverse=True)
                    
                    # Add to overall list
                    all_holdings.extend(fund_holdings)
                    successful_funds.append((fund_name, len(fund_holdings), total_value))
                    
                    print(f"   💰 Total Portfolio Value: ${total_value:,.0f}")
                    print(f"   🏆 Top Holding: {fund_holdings[0]['company']} (${fund_holdings[0]['value']:,.0f})")
                    
                else:
                    print(f"   📋 No holdings found")
            
            else:
                print(f"   ❌ Failed: {response.status_code}")
            
            # Rate limiting
            time.sleep(1)
            
        except Exception as e:
            print(f"   ❌ Error: {e}")
            continue
    
    return all_holdings, successful_funds

def analyze_combined_data(all_holdings):
    """Perform the three requested analyses on all funds combined"""
    
    if not all_holdings:
        print("No holdings data to analyze")
        return
    
    df = pd.DataFrame(all_holdings)
    
    print(f"\n" + "="*70)
    print("COMBINED 13F HOLDINGS ANALYSIS - Q1 2025")
    print("="*70)
    
    # 1. Top 25 most frequently held stocks
    print(f"\n1. TOP 25 MOST FREQUENTLY HELD STOCKS:")
    print("-" * 45)
    
    from collections import Counter
    stock_frequency = Counter()
    
    for _, holding in df.iterrows():
        key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        stock_frequency[key] += 1
    
    top_frequent = stock_frequency.most_common(25)
    for i, (stock, count) in enumerate(top_frequent, 1):
        print(f"{i:2d}. {stock}: {count} funds")
    
    # 2. Top holdings by total value
    print(f"\n\n2. TOP 25 HOLDINGS BY TOTAL VALUE:")
    print("-" * 45)
    
    from collections import defaultdict
    value_totals = defaultdict(float)
    
    for _, holding in df.iterrows():
        key = f"{holding['company']} ({holding['ticker']})" if holding['ticker'] else holding['company']
        value_totals[key] += holding['value']
    
    top_values = sorted(value_totals.items(), key=lambda x: x[1], reverse=True)[:25]
    for i, (stock, total_value) in enumerate(top_values, 1):
        print(f"{i:2d}. {stock}: ${total_value:,.0f}")
    
    # 3. Calculate percentage weights
    print(f"\n\n3. TOP 25 HOLDINGS BY WEIGHT (% of total portfolio value):")
    print("-" * 60)
    
    total_value_all = df['value'].sum()
    weight_analysis = []
    
    for stock, total_value in top_values[:25]:
        weight_percent = (total_value / total_value_all) * 100
        weight_analysis.append((stock, weight_percent, total_value))
    
    for i, (stock, weight, value) in enumerate(weight_analysis, 1):
        print(f"{i:2d}. {stock}: {weight:.2f}% (${value:,.0f})")
    
    # Summary statistics
    print(f"\n\n" + "="*70)
    print("SUMMARY STATISTICS")
    print("="*70)
    print(f"Total funds analyzed: {len(set(df['fund_name']))}")
    print(f"Total unique holdings: {len(set(df['company']))}")
    print(f"Total portfolio value: ${total_value_all:,.0f}")
    print(f"Average holding value: ${df['value'].mean():,.0f}")
    
    return df, top_frequent, top_values, weight_analysis

def create_excel_report(df, top_frequent, top_values, weight_analysis):
    """Create comprehensive Excel report"""
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: Most Frequently Held Stocks
    ws1 = wb.active
    ws1.title = "Most Frequent Holdings"
    
    headers = ["Rank", "Stock", "Number of Funds"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for i, (stock, count) in enumerate(top_frequent, 2):
        ws1.cell(row=i, column=1, value=i-1)
        ws1.cell(row=i, column=2, value=stock)
        ws1.cell(row=i, column=3, value=count)
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws1.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: Top Holdings by Value
    ws2 = wb.create_sheet("Top Holdings by Value")
    
    headers = ["Rank", "Stock", "Total Value ($)"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for i, (stock, value) in enumerate(top_values, 2):
        ws2.cell(row=i, column=1, value=i-1)
        ws2.cell(row=i, column=2, value=stock)
        ws2.cell(row=i, column=3, value=value)
    
    # Auto-adjust column widths
    for column in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws2.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 3: Raw Data
    ws3 = wb.create_sheet("Raw Data")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws3.append(r)
    
    # Style headers
    for cell in ws3[1:1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    wb.save('ALL_HEDGE_FUNDS_13F_Q1_2025.xlsx')

def main():
    print("ALL 16 HEDGE FUNDS 13F SCRAPER")
    print("="*40)
    
    # Get data for all funds
    all_holdings, successful_funds = get_all_hedge_funds_13f()
    
    print(f"\n" + "="*70)
    print("PROCESSING RESULTS")
    print("="*70)
    
    print(f"✅ Successfully processed {len(successful_funds)} funds:")
    for fund_name, holdings_count, total_value in successful_funds:
        print(f"   📊 {fund_name}: {holdings_count} holdings, ${total_value:,.0f}")
    
    if all_holdings:
        # Save raw data
        df = pd.DataFrame(all_holdings)
        df.to_csv('ALL_HEDGE_FUNDS_13F_Q1_2025.csv', index=False)
        
        # Perform analysis
        df, top_frequent, top_values, weight_analysis = analyze_combined_data(all_holdings)
        
        # Create Excel report
        create_excel_report(df, top_frequent, top_values, weight_analysis)
        
        print(f"\n🎉 FINAL FILES CREATED:")
        print(f"- ALL_HEDGE_FUNDS_13F_Q1_2025.csv")
        print(f"- ALL_HEDGE_FUNDS_13F_Q1_2025.xlsx")
        
    else:
        print(f"❌ No holdings data retrieved")

if __name__ == "__main__":
    main()