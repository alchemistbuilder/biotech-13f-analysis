import pandas as pd
import os
from collections import Counter, defaultdict
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import glob

class CSV13FAnalyzer:
    def __init__(self, csv_directory="./"):
        self.csv_directory = csv_directory
        self.holdings_data = []
        
    def find_csv_files(self):
        """Find all 13F CSV and Excel files in the directory"""
        patterns = [
            "*.csv",
            "*.xlsx",
            "*.xls",
            "*13F*.csv", 
            "*13F*.xlsx",
            "*Holdings*.csv",
            "*Holdings*.xlsx",
            "*Portfolio*.csv",
            "*Portfolio*.xlsx"
        ]
        
        csv_files = []
        for pattern in patterns:
            csv_files.extend(glob.glob(os.path.join(self.csv_directory, pattern)))
        
        # Remove duplicates
        csv_files = list(set(csv_files))
        
        print(f"Found {len(csv_files)} CSV files:")
        for file in csv_files:
            print(f"  - {os.path.basename(file)}")
            
        return csv_files
    
    def parse_csv_file(self, csv_file):
        """Parse a single 13F CSV file"""
        try:
            # Extract fund name from filename
            filename = os.path.basename(csv_file)
            fund_name = filename.replace(".csv", "").replace(".xlsx", "").replace(".xls", "").replace("_", " ").title()
            
            # Try different common file formats
            if csv_file.endswith('.xlsx') or csv_file.endswith('.xls'):
                df = pd.read_excel(csv_file)
            else:
                df = pd.read_csv(csv_file)
            
            print(f"Processing {fund_name}...")
            print(f"  Columns: {list(df.columns)}")
            print(f"  Rows: {len(df)}")
            
            # Common column mappings (case insensitive)
            column_mappings = {
                'company': ['company', 'issuer', 'issuer name', 'security name', 'name'],
                'ticker': ['ticker', 'symbol', 'sym', 'stock symbol', 'security symbol'],
                'shares': ['shares', 'quantity', 'shares held', 'position', 'shares owned'],
                'value': ['value', 'market value', 'fair value', 'position value', 'market cap', 'value ($000)', 'value (000s)'],
                'weight': ['weight', 'percent', '%', 'portfolio %', '% of portfolio', 'allocation']
            }
            
            # Find matching columns
            matched_columns = {}
            df_columns_lower = [col.lower() for col in df.columns]
            
            for field, possible_names in column_mappings.items():
                for possible_name in possible_names:
                    if possible_name.lower() in df_columns_lower:
                        actual_column = df.columns[df_columns_lower.index(possible_name.lower())]
                        matched_columns[field] = actual_column
                        break
            
            print(f"  Matched columns: {matched_columns}")
            
            # Process each row
            holdings_count = 0
            for _, row in df.iterrows():
                try:
                    # Extract data using matched columns
                    company = row.get(matched_columns.get('company', ''), '').strip()
                    ticker = row.get(matched_columns.get('ticker', ''), '').strip()
                    
                    # Handle value field
                    value_raw = row.get(matched_columns.get('value', ''), 0)
                    if isinstance(value_raw, str):
                        value_raw = value_raw.replace('$', '').replace(',', '').replace('(', '').replace(')', '')
                        # Handle thousands notation
                        if 'K' in value_raw.upper():
                            value_raw = value_raw.upper().replace('K', '')
                            multiplier = 1000
                        elif 'M' in value_raw.upper():
                            value_raw = value_raw.upper().replace('M', '')
                            multiplier = 1000000
                        elif 'B' in value_raw.upper():
                            value_raw = value_raw.upper().replace('B', '')
                            multiplier = 1000000000
                        else:
                            multiplier = 1
                        
                        try:
                            value = float(value_raw) * multiplier
                        except:
                            value = 0
                    else:
                        value = float(value_raw) if pd.notna(value_raw) else 0
                    
                    # Handle shares
                    shares_raw = row.get(matched_columns.get('shares', ''), 0)
                    if isinstance(shares_raw, str):
                        shares_raw = shares_raw.replace(',', '')
                        try:
                            shares = float(shares_raw)
                        except:
                            shares = 0
                    else:
                        shares = float(shares_raw) if pd.notna(shares_raw) else 0
                    
                    # Add to holdings if we have valid data
                    if company and (value > 0 or shares > 0):
                        self.holdings_data.append({
                            'fund_name': fund_name,
                            'company': company,
                            'ticker': ticker,
                            'shares': shares,
                            'value': value
                        })
                        holdings_count += 1
                        
                except Exception as e:
                    continue
            
            print(f"  Successfully parsed {holdings_count} holdings")
            return holdings_count
            
        except Exception as e:
            print(f"Error processing {csv_file}: {e}")
            return 0
    
    def process_all_csv_files(self):
        """Process all CSV files in the directory"""
        csv_files = self.find_csv_files()
        
        if not csv_files:
            print("No CSV files found. Please place your 13F CSV files in the current directory.")
            return False
        
        total_processed = 0
        for csv_file in csv_files:
            processed = self.parse_csv_file(csv_file)
            total_processed += processed
        
        print(f"\nTotal holdings processed: {len(self.holdings_data)}")
        print(f"From {len(set(h['fund_name'] for h in self.holdings_data))} different funds")
        
        return len(self.holdings_data) > 0
    
    def analyze_holdings(self):
        """Perform the requested analysis"""
        if not self.holdings_data:
            print("No holdings data to analyze")
            return None, None, None, None
            
        df = pd.DataFrame(self.holdings_data)
        
        print("\n" + "="*60)
        print("13F HOLDINGS ANALYSIS")
        print("="*60)
        
        # 1. Top 25 most frequently held stocks
        print("\n1. TOP 25 MOST FREQUENTLY HELD STOCKS:")
        print("-" * 40)
        
        stock_frequency = Counter()
        for _, holding in df.iterrows():
            # Use company name primarily, add ticker if available
            if holding['ticker']:
                key = f"{holding['company']} ({holding['ticker']})"
            else:
                key = holding['company']
            stock_frequency[key] += 1
            
        top_frequent = stock_frequency.most_common(25)
        for i, (stock, count) in enumerate(top_frequent, 1):
            print(f"{i:2d}. {stock}: {count} funds")
        
        # 2. Top holdings by total value
        print("\n\n2. TOP 25 HOLDINGS BY TOTAL VALUE:")
        print("-" * 40)
        
        value_totals = defaultdict(float)
        for _, holding in df.iterrows():
            if holding['ticker']:
                key = f"{holding['company']} ({holding['ticker']})"
            else:
                key = holding['company']
            value_totals[key] += holding['value']
            
        top_values = sorted(value_totals.items(), key=lambda x: x[1], reverse=True)[:25]
        for i, (stock, total_value) in enumerate(top_values, 1):
            print(f"{i:2d}. {stock}: ${total_value:,.0f}")
        
        # 3. Calculate percentage weights
        print("\n\n3. TOP 25 HOLDINGS BY WEIGHT (% of total portfolio value):")
        print("-" * 50)
        
        total_value_all = df['value'].sum()
        weight_analysis = []
        
        for stock, total_value in top_values[:25]:
            weight_percent = (total_value / total_value_all) * 100 if total_value_all > 0 else 0
            weight_analysis.append((stock, weight_percent, total_value))
        
        for i, (stock, weight, value) in enumerate(weight_analysis, 1):
            print(f"{i:2d}. {stock}: {weight:.2f}% (${value:,.0f})")
        
        # Summary statistics
        print("\n\n" + "="*60)
        print("SUMMARY STATISTICS")
        print("="*60)
        unique_funds = len(set(df['fund_name']))
        unique_holdings = len(set(df['company']))
        print(f"Total funds analyzed: {unique_funds}")
        print(f"Total unique holdings: {unique_holdings}")
        print(f"Total portfolio value: ${total_value_all:,.0f}")
        if len(df) > 0:
            print(f"Average holding value: ${df['value'].mean():,.0f}")
        
        return df, top_frequent, top_values, weight_analysis
    
    def create_excel_report(self, df, top_frequent, top_values, weight_analysis):
        """Create an Excel report with all analysis results"""
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Sheet 1: Most Frequently Held Stocks
        ws1 = wb.active
        ws1.title = "Most Frequent Holdings"
        
        # Headers
        ws1['A1'] = "Rank"
        ws1['B1'] = "Stock"
        ws1['C1'] = "Number of Funds"
        
        # Style headers
        for cell in ws1[1:1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for i, (stock, count) in enumerate(top_frequent, 2):
            ws1[f'A{i}'] = i - 1
            ws1[f'B{i}'] = stock
            ws1[f'C{i}'] = count
        
        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Top Holdings by Value
        ws2 = wb.create_sheet("Top Holdings by Value")
        
        # Headers
        ws2['A1'] = "Rank"
        ws2['B1'] = "Stock"
        ws2['C1'] = "Total Value ($)"
        
        # Style headers
        for cell in ws2[1:1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for i, (stock, value) in enumerate(top_values, 2):
            ws2[f'A{i}'] = i - 1
            ws2[f'B{i}'] = stock
            ws2[f'C{i}'] = value
        
        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Top Holdings by Weight
        ws3 = wb.create_sheet("Top Holdings by Weight")
        
        # Headers
        ws3['A1'] = "Rank"
        ws3['B1'] = "Stock"
        ws3['C1'] = "Weight (%)"
        ws3['D1'] = "Total Value ($)"
        
        # Style headers
        for cell in ws3[1:1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for i, (stock, weight, value) in enumerate(weight_analysis, 2):
            ws3[f'A{i}'] = i - 1
            ws3[f'B{i}'] = stock
            ws3[f'C{i}'] = weight
            ws3[f'D{i}'] = value
        
        # Auto-adjust column widths
        for column in ws3.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws3.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 4: Raw Data
        ws4 = wb.create_sheet("Raw Data")
        
        # Add dataframe to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws4.append(r)
        
        # Style headers
        for cell in ws4[1:1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws4.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws4.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 5: Summary Statistics
        ws5 = wb.create_sheet("Summary")
        
        total_funds = len(set(df['fund_name']))
        total_holdings = len(set(df['company']))
        total_value = df['value'].sum()
        avg_value = df['value'].mean()
        
        ws5['A1'] = "13F Holdings Analysis Summary"
        ws5['A1'].font = Font(bold=True, size=16)
        
        ws5['A3'] = "Total funds analyzed:"
        ws5['B3'] = total_funds
        
        ws5['A4'] = "Total unique holdings:"
        ws5['B4'] = total_holdings
        
        ws5['A5'] = "Total portfolio value:"
        ws5['B5'] = total_value
        
        ws5['A6'] = "Average holding value:"
        ws5['B6'] = avg_value
        
        ws5['A8'] = "Fund List:"
        ws5['A8'].font = Font(bold=True)
        
        fund_names = sorted(set(df['fund_name']))
        for i, fund in enumerate(fund_names, 9):
            ws5[f'A{i}'] = fund
        
        # Auto-adjust column widths
        for column in ws5.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws5.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save('13f_analysis_report.xlsx')
        print(f"\nExcel report saved as: 13f_analysis_report.xlsx")

def main():
    print("13F Holdings CSV Analyzer")
    print("========================")
    print("This script will analyze CSV files containing 13F holdings data.")
    print("Please place your CSV files in the current directory.\n")
    
    analyzer = CSV13FAnalyzer()
    
    # Process CSV files
    if analyzer.process_all_csv_files():
        # Perform analysis
        df, top_frequent, top_values, weight_analysis = analyzer.analyze_holdings()
        
        if df is not None:
            # Create Excel report
            analyzer.create_excel_report(df, top_frequent, top_values, weight_analysis)
            
            # Also save CSV and JSON
            df.to_csv('13f_holdings_raw_data.csv', index=False)
            
            analysis_results = {
                'most_frequent_stocks': top_frequent,
                'top_value_holdings': top_values,
                'top_weight_holdings': weight_analysis
            }
            
            with open('13f_analysis_results.json', 'w') as f:
                json.dump(analysis_results, f, indent=2, default=str)
            
            print(f"\nAdditional files saved:")
            print(f"- Raw data: 13f_holdings_raw_data.csv")
            print(f"- Analysis: 13f_analysis_results.json")
        
    else:
        print("\nNo CSV files were successfully processed.")
        print("\nTo use this analyzer:")
        print("1. Place your 13F CSV files in this directory")
        print("2. Run the script again")
        print("\nExpected CSV columns (any combination):")
        print("- Company/Issuer Name")
        print("- Ticker/Symbol") 
        print("- Shares/Quantity")
        print("- Value/Market Value")

if __name__ == "__main__":
    main()